"""Dump every formula and value from sheet/shop_titans.xlsx as per-tab TSVs.

Loads the workbook twice: once preserving formulas (data_only=False) and once
with cached calc results (data_only=True). For each non-empty cell we emit the
formula or the value, plus the cached evaluation if it's a formula.
"""
import os
import openpyxl

HERE = os.path.dirname(__file__)
SRC = os.path.join(HERE, "shop_titans.xlsx")
OUT_DIR = os.path.join(HERE, "formulas")
os.makedirs(OUT_DIR, exist_ok=True)

wb_f = openpyxl.load_workbook(SRC, data_only=False)
wb_v = openpyxl.load_workbook(SRC, data_only=True)

with open(os.path.join(OUT_DIR, "_defined_names.tsv"), "w", encoding="utf-8") as f:
    f.write("name\tvalue\n")
    for name in wb_f.defined_names:
        d = wb_f.defined_names[name]
        f.write(f"{name}\t{d.value}\n")

index = []
for sheet_name in wb_f.sheetnames:
    ws_f = wb_f[sheet_name]
    ws_v = wb_v[sheet_name]
    safe = "".join(c if c.isalnum() or c in "-_ " else "_" for c in sheet_name).strip()
    out_path = os.path.join(OUT_DIR, f"{safe}.tsv")
    fc = vc = 0
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(f"# Sheet: {sheet_name}\n")
        f.write(f"# Dimensions: {ws_f.dimensions}, rows={ws_f.max_row}, cols={ws_f.max_column}\n")
        f.write("cell\ttype\tcontent\tcached\n")
        for row in ws_f.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                ref = cell.coordinate
                v = cell.value
                if isinstance(v, str) and v.startswith("="):
                    cached = ws_v[ref].value
                    fc += 1
                    sv = v.replace("\t", " ").replace("\n", " ")
                    sc = str(cached).replace("\t", " ").replace("\n", " ") if cached is not None else ""
                    f.write(f"{ref}\tF\t{sv}\t{sc}\n")
                else:
                    vc += 1
                    sv = str(v).replace("\t", " ").replace("\n", " ")
                    f.write(f"{ref}\tV\t{sv}\t\n")
    index.append((sheet_name, fc, vc))
    print(f"{sheet_name}: {fc} formulas, {vc} values")

with open(os.path.join(OUT_DIR, "_index.tsv"), "w", encoding="utf-8") as f:
    f.write("sheet\tformulas\tvalues\n")
    for s, fc, vc in index:
        f.write(f"{s}\t{fc}\t{vc}\n")
